from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
wb=load_workbook(r"C:\Users\ashis\OneDrive\Desktop\Data Automation\excell_automaton\practice.xlsx")
ws=wb.active
ws.merge_cells("a1:b1")  # ("  :  ") cell no. between merged 
wb.save(r"C:\Users\ashis\OneDrive\Desktop\Data Automation\excell_automaton\practice.xlsx")