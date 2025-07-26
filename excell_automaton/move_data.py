# move the range of data from one place to another 
from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
wb=load_workbook(r"C:\Users\ashis\OneDrive\Desktop\Data Automation\excell_automaton\practice.xlsx")
ws=wb.active
ws.move_range ("g5:k7",rows=-3,cols=-6)   # for row 1=downward -1=upward for column 1=right side and -1=left side 
wb.save(r"C:\Users\ashis\OneDrive\Desktop\Data Automation\excell_automaton\practice.xlsx")