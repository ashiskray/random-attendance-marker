from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
wb=load_workbook(r"C:\Users\ashis\OneDrive\Desktop\Data Automation\collagedata.xlsx")
ws=wb.active
for r in range (1,5):
    for c in range(1,7):
        cell=ws.cell(r,c)
        print(str(cell)+" : "+ str (cell.value))