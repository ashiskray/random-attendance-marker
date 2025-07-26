from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
wb=load_workbook(r"C:\Users\ashis\OneDrive\Desktop\Data Automation\excell_automaton\practice.xlsx")
ws=wb.active
ws.unmerge_cells("a1:b1")
wb.save(r"C:\Users\ashis\OneDrive\Desktop\Data Automation\excell_automaton\practice.xlsx")
#  but unmerging cell will delete the data of merged cell permanently 