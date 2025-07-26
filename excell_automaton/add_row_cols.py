# add row and column in sheet 
from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
wb=load_workbook(r"C:\Users\ashis\OneDrive\Desktop\Data Automation\excell_automaton\practice.xlsx")
ws=wb.active
ws.insert_rows(3) # row no. where to insert 
ws.insert_cols(3) # cols no. where to insert
wb.save(r"C:\Users\ashis\OneDrive\Desktop\Data Automation\excell_automaton\practice.xlsx")