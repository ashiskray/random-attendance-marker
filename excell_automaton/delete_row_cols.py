# to dellete perticular row and column 
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
wb=load_workbook(r"C:\Users\ashis\OneDrive\Desktop\Data Automation\excell_automaton\practice.xlsx")
ws=wb.active
ws.delete_rows(3)
ws.delete_cols(3)
wb.save(r"C:\Users\ashis\OneDrive\Desktop\Data Automation\excell_automaton\practice.xlsx")